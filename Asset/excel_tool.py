import os
import pandas as pd
from pathlib import Path
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font

class ExcelTool:
    @staticmethod
    def main_menu(folder_path=None):
        """Main menu for merging Excel files with two options."""
        while True:
            print("\nSelect an option:")
            print("1. Merge Workbook (Separate sheets)")
            print("2. Merge Worksheet (Single sheet)")
            print("0. Exit to Main Menu")
            
            choice = input("Enter the number of the option: ").strip()

            if choice == '1':
                if folder_path:
                    ExcelTool.merge_workbook(folder_path)
                else:
                    print("No folder selected. Please select a folder first.")

            elif choice == '2':
                if folder_path:
                    ExcelTool.merge_worksheet(folder_path)
                else:
                    print("No folder selected. Please select a folder first.")

            elif choice == '0':
                break  # Exit back to the main menu

            else:
                print("Invalid choice. Please enter a valid option.")

    @staticmethod
    def merge_workbook(folder_path):
        try:
            output_folder = Path('Output')
            output_folder.mkdir(parents=True, exist_ok=True)
            output_path = output_folder / "merged_workbook.xlsx"

            with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                for idx, file_name in enumerate(sorted(os.listdir(folder_path))):
                    if file_name.endswith(('.xlsx', '.csv')):
                        file_path = Path(folder_path) / file_name
                        sheet_name = file_name[:31] if len(file_name) > 31 else file_name
                        
                        df = pd.read_csv(file_path) if file_name.endswith('.csv') else pd.read_excel(file_path)
                        df['Source File'] = file_name
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        logging.info(f"Added {file_name} as sheet {sheet_name}")

            return True
        except Exception as e:
            logging.error(f"Error in merging workbook: {e}")
            raise

    @staticmethod
    def merge_worksheet(folder_path):
        try:
            output_folder = Path('Output')
            output_folder.mkdir(parents=True, exist_ok=True)
            output_path = output_folder / "merged_worksheet.xlsx"
            
            all_dfs = []
            for file_name in sorted(os.listdir(folder_path)):
                if file_name.endswith(('.xlsx', '.csv')):
                    file_path = Path(folder_path) / file_name
                    df = pd.read_csv(file_path) if file_name.endswith('.csv') else pd.read_excel(file_path)
                    df['Source File'] = file_name
                    all_dfs.append(df)

            if all_dfs:
                combined_df = pd.concat(all_dfs, ignore_index=True)
                combined_df.to_excel(output_path, index=False)
                logging.info(f"All files merged into single worksheet at {output_path}")
                return True
            else:
                raise ValueError("No Excel or CSV files found in the specified folder")
        except Exception as e:
            logging.error(f"Error in merging worksheet: {e}")
            raise

    @staticmethod
    def excel_to_csv(excel_path):
        try:
            output_folder = Path('Output')
            output_folder.mkdir(parents=True, exist_ok=True)
            
            xls = pd.ExcelFile(excel_path)
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                output_path = output_folder / f"{Path(excel_path).stem}_{sheet_name}.csv"
                df.to_csv(output_path, index=False)
                logging.info(f"Sheet {sheet_name} converted to CSV: {output_path}")
            
            return True
        except Exception as e:
            logging.error(f"Error converting Excel to CSV: {e}")
            raise

    @staticmethod
    def csv_to_excel(csv_path):
        try:
            output_folder = Path('Output')
            output_folder.mkdir(parents=True, exist_ok=True)
            
            df = pd.read_csv(csv_path)
            output_path = output_folder / f"{Path(csv_path).stem}.xlsx"
            
            # Create a styled Excel file
            writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Get workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Add some formatting
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D7E4BC',
                'border': 1
            })
            
            # Format the header row
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Auto-adjust columns width
            for idx, col in enumerate(df.columns):
                series = df[col]
                max_len = max(
                    series.astype(str).map(len).max(),
                    len(str(series.name))
                ) + 1
                worksheet.set_column(idx, idx, max_len)
            
            writer.close()
            logging.info(f"CSV converted to Excel: {output_path}")
            return True
        except Exception as e:
            logging.error(f"Error converting CSV to Excel: {e}")
            raise

    @staticmethod
    def format_excel(excel_path):
        try:
            output_folder = Path('Output')
            output_folder.mkdir(parents=True, exist_ok=True)
            
            wb = openpyxl.load_workbook(excel_path)
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # Format header row
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                header_font = Font(bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            output_path = output_folder / f"{Path(excel_path).stem}_formatted.xlsx"
            wb.save(output_path)
            logging.info(f"Excel file formatted and saved to {output_path}")
            return True
        except Exception as e:
            logging.error(f"Error formatting Excel file: {e}")
            raise
